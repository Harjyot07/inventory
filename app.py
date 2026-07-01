from flask import Flask, render_template_string, request, redirect, send_file
from datetime import datetime
import csv, io, os

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    openpyxl = None
try:
    from fpdf import FPDF
except ImportError:
    FPDF = None

app = Flask(__name__)

INBOUND_FILE = "inbound.csv"
OUTBOUND_FILE = "outbound.csv"

INBOUND_HEADERS = ["DATE","PRODUCT","SPECIFICATION","NO OF UNIT","REMARKS"]
OUTBOUND_HEADERS = ["DATE","PRODUCT","DEPARTMENT","EMPLOYEE NAME","QUANTITY","REMARKS"]

def load_csv(path, headers):
    rows = []
    if os.path.exists(path):
        with open(path, newline="") as f:
            reader = csv.DictReader(f)
            for r in reader:
                rows.append(r)
    return rows

def save_csv(path, headers, rows):
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)

def load_inbound():
    return load_csv(INBOUND_FILE, INBOUND_HEADERS)

def save_inbound(rows):
    save_csv(INBOUND_FILE, INBOUND_HEADERS, rows)

def load_outbound():
    return load_csv(OUTBOUND_FILE, OUTBOUND_HEADERS)

def save_outbound(rows):
    save_csv(OUTBOUND_FILE, OUTBOUND_HEADERS, rows)

HTML = """<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>QUANTAM-JWIZ-INVENTORY</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Inter,-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;background:#0f172a;color:#e2e8f0}
.header{background:linear-gradient(135deg,#1e293b 0%,#0f172a 100%);border-bottom:1px solid #334155;padding:20px 0 16px;position:sticky;top:0;z-index:10}
.header h1{font-size:22px;font-weight:700;color:#f1f5f9;letter-spacing:-0.3px}
.header .sub{font-size:13px;color:#94a3b8;margin-top:4px}
.container{max-width:1280px;margin:0 auto;padding:0 24px}
.content{padding:24px 0}
.alert{padding:12px 20px;border-radius:10px;margin-bottom:20px;font-size:14px;font-weight:500}
.alert-success{background:#064e3b;color:#6ee7b7;border:1px solid #065f46}
.stats{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:16px;margin-bottom:24px}
.stat-card{background:#1e293b;border-radius:12px;padding:18px 22px;border:1px solid #334155}
.stat-card .stat-label{font-size:12px;font-weight:600;color:#64748b;text-transform:uppercase;letter-spacing:0.5px}
.stat-card .stat-value{font-size:28px;font-weight:700;color:#f1f5f9;margin-top:4px}
.stat-card .stat-sub{font-size:12px;color:#94a3b8;margin-top:2px}
.stat-card.purple .stat-value{color:#a78bfa}
.stat-card.green .stat-value{color:#34d399}
.stat-card.blue .stat-value{color:#60a5fa}
.stat-card.amber .stat-value{color:#fbbf24}
.stat-card.pink .stat-value{color:#f472b6}
.stat-card.indigo .stat-value{color:#818cf8}
.card{background:#1e293b;border-radius:12px;padding:24px;margin-bottom:20px;border:1px solid #334155}
.card h2{font-size:16px;font-weight:600;color:#f1f5f9;margin-bottom:16px}
.tabs{display:flex;gap:0;margin-bottom:20px;border-bottom:1px solid #334155}
.tab{padding:10px 24px;cursor:pointer;font-weight:600;background:transparent;color:#64748b;border:none;font-size:14px;transition:.2s;border-bottom:2px solid transparent;margin-bottom:-1px}
.tab.active{color:#60a5fa;border-bottom-color:#60a5fa}
.tab:hover{color:#94a3b8}
.form-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:14px}
.form-group{display:flex;flex-direction:column}
.form-group label{font-size:12px;font-weight:600;margin-bottom:6px;color:#94a3b8;text-transform:uppercase;letter-spacing:0.3px}
.form-group input,.form-group select{padding:10px 14px;border:1px solid #334155;border-radius:8px;font-size:14px;outline:0;transition:.2s;background:#0f172a;color:#e2e8f0}
.form-group input:focus,.form-group select:focus{border-color:#60a5fa;box-shadow:0 0 0 3px rgba(96,165,250,.15)}
.form-group input::placeholder{color:#475569}
.btn{display:inline-flex;align-items:center;gap:6px;padding:10px 22px;border:none;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;transition:.2s;text-decoration:none}
.btn-primary{background:#3b82f6;color:#fff}
.btn-primary:hover{background:#2563eb}
.btn-success{background:#059669;color:#fff}
.btn-success:hover{background:#047857}
.btn-warning{background:#d97706;color:#fff}
.btn-warning:hover{background:#b45309}
.btn-info{background:#0891b2;color:#fff;padding:6px 14px;font-size:12px}
.btn-info:hover{background:#0e7490}
.btn-danger{background:#dc2626;color:#fff;padding:6px 14px;font-size:12px}
.btn-danger:hover{background:#b91c1c}
.btn-sm{padding:6px 14px;font-size:12px}
.actions{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:10px}
.actions .btn{font-size:12px;padding:8px 16px}
table{width:100%;border-collapse:collapse;font-size:13px}
th,td{text-align:left;padding:10px 12px;border-bottom:1px solid #1e293b}
th{background:#0f172a;font-weight:600;color:#94a3b8;position:sticky;top:0}
tr:hover{background:#0f172a}
.table-wrap{overflow-x:auto;max-height:520px;overflow-y:auto;border-radius:8px;border:1px solid #334155}
.table-wrap table{border-radius:8px}
.empty{text-align:center;padding:40px 20px;color:#475569;font-size:14px}
.footer{text-align:center;color:#475569;font-size:12px;padding:24px 0 12px;border-top:1px solid #1e293b;margin-top:8px}
.modal-overlay{display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,.6);z-index:100;align-items:center;justify-content:center}
.modal-overlay.open{display:flex}
.modal-content{background:#1e293b;border-radius:14px;border:1px solid #334155;width:90%;max-width:520px;max-height:80vh;overflow-y:auto;padding:0}
.modal-header{display:flex;justify-content:space-between;align-items:center;padding:18px 24px;border-bottom:1px solid #334155}
.modal-header h3{font-size:16px;font-weight:600;color:#f1f5f9;margin:0}
.modal-close{background:none;border:none;color:#64748b;font-size:24px;cursor:pointer;padding:0;line-height:1}
.modal-close:hover{color:#f1f5f9}
.modal-body{padding:20px 24px}
.detail-table{width:100%;border-collapse:collapse}
.detail-table tr{border-bottom:1px solid #334155}
.detail-table tr:last-child{border-bottom:none}
.detail-label{width:40%;padding:10px 12px 10px 0;font-size:12px;font-weight:600;color:#64748b;text-transform:uppercase;letter-spacing:0.3px;vertical-align:top}
.detail-value{width:60%;padding:10px 0;font-size:14px;color:#e2e8f0;word-break:break-word}
::-webkit-scrollbar{width:6px;height:6px}
::-webkit-scrollbar-track{background:#0f172a}
::-webkit-scrollbar-thumb{background:#334155;border-radius:3px}
::-webkit-scrollbar-thumb:hover{background:#475569}
@media(max-width:600px){.form-grid{grid-template-columns:1fr}.stats{grid-template-columns:1fr 1fr}}
</style>
</head>
<body>
<div class="header">
<div class="container">
<h1>QUANTAM-JWIZ-INVENTORY</h1>
<p class="sub">Inbound &mdash; New Stock Arrival &nbsp;|&nbsp; Outbound &mdash; Issued to Person</p>
</div>
</div>
<div class="content">
<div class="container">

{% if merged %}
<div class="alert alert-success">Merged with existing inbound entry &mdash; quantity updated!</div>
{% endif %}

<div class="stats">
<div class="stat-card purple">
<div class="stat-label">Total Inbound Stock</div>
<div class="stat-value">{{ total_qty }}</div>
<div class="stat-sub">{{ inbound|length }} item types</div>
</div>
<div class="stat-card green">
<div class="stat-label">Items Issued</div>
<div class="stat-value">{{ outbound|length }}</div>
<div class="stat-sub">outbound entries</div>
</div>
<div class="stat-card blue">
<div class="stat-label">Unique Persons</div>
<div class="stat-value">{{ unique_persons }}</div>
<div class="stat-sub">receivers</div>
</div>
<div class="stat-card amber">
<div class="stat-label">Total Entries</div>
<div class="stat-value">{{ inbound|length + outbound|length }}</div>
<div class="stat-sub">combined records</div>
</div>
<div class="stat-card pink">
<div class="stat-label">PC Issued</div>
<div class="stat-value">{{ pc_count }}</div>
<div class="stat-sub">units issued</div>
</div>
<div class="stat-card amber">
<div class="stat-label">Keyboard Issued</div>
<div class="stat-value">{{ keyboard_count }}</div>
<div class="stat-sub">units issued</div>
</div>
<div class="stat-card green">
<div class="stat-label">Mouse Issued</div>
<div class="stat-value">{{ mouse_count }}</div>
<div class="stat-sub">units issued</div>
</div>
{% for dept, qty in dept_items %}
<div class="stat-card indigo">
<div class="stat-label">{{ dept }}</div>
<div class="stat-value">{{ qty }}</div>
<div class="stat-sub">items issued</div>
</div>
{% endfor %}
</div>

<div class="tabs">
<button class="tab active" data-tab="inbound">Inbound (Stock In)</button>
<button class="tab" data-tab="outbound">Outbound (Issued to Person)</button>
</div>

<div id="tab-inbound" class="tab-content">
<div class="card">
<h2>Add Inbound Stock</h2>
<form method="POST" action="/add/inbound">
<div class="form-grid">
{% for field in inbound_fields %}
{% if field != "NO OF UNIT" and field != "DATE" %}
<div class="form-group">
<label for="in_{{ field }}">{{ field }}</label>
<input type="text" id="in_{{ field }}" name="{{ field }}" placeholder="Enter {{ field|lower }}" required>
</div>
{% endif %}
{% endfor %}
<div class="form-group">
<label for="in_NOOFUNIT">NO OF UNIT</label>
<input type="number" id="in_NOOFUNIT" name="NO OF UNIT" placeholder="Units" required min="1">
</div>
</div>
<div style="margin-top:16px">
<button type="submit" class="btn btn-primary">Add Inbound Stock</button>
</div>
</form>
</div>

<div class="card">
<h2>Import Inbound from Excel</h2>
<form method="POST" action="/import/inbound" enctype="multipart/form-data">
<div style="display:flex;gap:12px;align-items:flex-end;flex-wrap:wrap">
<div class="form-group" style="flex:1;min-width:200px">
<label>Choose .xlsx file</label>
<input type="file" name="file" accept=".xlsx" required>
</div>
<div class="form-group" style="min-width:120px">
<label>Sheet name</label>
<input type="text" name="sheet" placeholder="leave blank for active sheet">
</div>
<button type="submit" class="btn btn-success">Upload &amp; Import</button>
</div>
</form>
</div>

<div class="card">
<div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px;margin-bottom:16px">
<h2 style="margin:0">Inbound Stock List ({{ inbound|length }})</h2>
<div class="actions">
<a href="/export/csv" class="btn btn-success">CSV</a>
<a href="/export/excel" class="btn btn-success">Excel</a>
<a href="/export/pdf" class="btn btn-warning">PDF</a>
</div>
</div>
{% if inbound %}
<div class="table-wrap">
<table>
<thead><tr><th>S.NO</th>{% for field in inbound_fields %}<th>{{ field }}</th>{% endfor %}<th>Action</th></tr></thead>
<tbody>
{% for row in inbound %}
<tr>
<td>{{ loop.index }}</td>
{% for field in inbound_fields %}<td>{{ row[field] }}</td>{% endfor %}
<td>
<button onclick='viewEntry("inbound", {{ loop.index0 }})' class="btn btn-info btn-sm">View</button>
<a href="/edit/inbound/{{ loop.index0 }}" class="btn btn-warning btn-sm">Edit</a>
<a href="/delete/inbound/{{ loop.index0 }}" class="btn btn-danger" onclick="return confirm('Delete this entry?')">Delete</a>
</td>
</tr>
{% endfor %}
</tbody>
</table>
</div>
{% else %}
<div class="empty">No inbound stock entries yet.</div>
{% endif %}
</div>
</div>

<div id="tab-outbound" class="tab-content" style="display:none">
<div class="card">
<h2>Assign Item to Person (Outbound)</h2>
<form method="POST" action="/add/outbound">
<div class="form-grid">
{% for field in outbound_fields %}
{% if field != "DATE" %}
<div class="form-group">
<label for="out_{{ field }}">{{ field }}</label>
<input type="text" id="out_{{ field }}" name="{{ field }}" placeholder="Enter {{ field|lower }}" required>
</div>
{% endif %}
{% endfor %}
</div>
<div style="margin-top:16px">
<button type="submit" class="btn btn-primary">Assign Item</button>
</div>
</form>
</div>

<div class="card">
<h2>Import Outbound from Excel</h2>
<form method="POST" action="/import/outbound" enctype="multipart/form-data">
<div style="display:flex;gap:12px;align-items:flex-end;flex-wrap:wrap">
<div class="form-group" style="flex:1;min-width:200px">
<label>Choose .xlsx file</label>
<input type="file" name="file" accept=".xlsx" required>
</div>
<div class="form-group" style="min-width:120px">
<label>Sheet name</label>
<input type="text" name="sheet" placeholder="leave blank for active sheet">
</div>
<button type="submit" class="btn btn-success">Upload &amp; Import</button>
</div>
</form>
</div>

<div class="card">
<div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px;margin-bottom:16px">
<h2 style="margin:0">Outbound List ({{ outbound|length }})</h2>
<div class="actions">
<a href="/export/csv" class="btn btn-success">CSV</a>
<a href="/export/excel" class="btn btn-success">Excel</a>
<a href="/export/pdf" class="btn btn-warning">PDF</a>
</div>
</div>
{% if outbound %}
<div class="table-wrap">
<table>
<thead><tr><th>S.NO</th>{% for field in outbound_fields %}<th>{{ field }}</th>{% endfor %}<th>Action</th></tr></thead>
<tbody>
{% for row in outbound %}
<tr>
<td>{{ loop.index }}</td>
{% for field in outbound_fields %}<td>{{ row[field] }}</td>{% endfor %}
<td>
<button onclick='viewEntry("outbound", {{ loop.index0 }})' class="btn btn-info btn-sm">View</button>
<a href="/edit/outbound/{{ loop.index0 }}" class="btn btn-warning btn-sm">Edit</a>
<a href="/delete/outbound/{{ loop.index0 }}" class="btn btn-danger" onclick="return confirm('Delete this entry?')">Delete</a>
</td>
</tr>
{% endfor %}
</tbody>
</table>
</div>
{% else %}
<div class="empty">No outbound entries yet.</div>
{% endif %}
</div>
</div>

<div id="viewModal" class="modal-overlay" onclick="closeModal(event)">
<div class="modal-content" onclick="event.stopPropagation()">
<div class="modal-header">
<h3 id="modalTitle">Entry Details</h3>
<button class="modal-close" onclick="closeModal()">&times;</button>
</div>
<div class="modal-body" id="modalBody"></div>
</div>
</div>

<div class="footer">QUANTAM-JWIZ-INVENTORY &mdash; {{ now }}</div>
</div>

<script>
var inboundData = {{ inbound | tojson }};
var outboundData = {{ outbound | tojson }};
var inboundFields = {{ inbound_fields | tojson }};
var outboundFields = {{ outbound_fields | tojson }};

function viewEntry(type, idx) {
var data = type === 'inbound' ? inboundData : outboundData;
var fields = type === 'inbound' ? inboundFields : outboundFields;
var row = data[idx];
if (!row) return;
var title = (type === 'inbound' ? 'Inbound' : 'Outbound') + ' Entry #' + (idx + 1);
document.getElementById('modalTitle').textContent = title;
var html = '<table class="detail-table">';
fields.forEach(function(f) {
html += '<tr><td class="detail-label">' + f + '</td><td class="detail-value">' + (row[f] || '') + '</td></tr>';
});
html += '</table>';
document.getElementById('modalBody').innerHTML = html;
document.getElementById('viewModal').classList.add('open');
}

function closeModal(e) {
if (!e || !e.target || e.target === document.getElementById('viewModal')) {
document.getElementById('viewModal').classList.remove('open');
}
}

document.querySelectorAll('.tab').forEach(function(btn) {
btn.addEventListener('click', function() {
var name = this.getAttribute('data-tab');
document.querySelectorAll('.tab-content').forEach(function(el) { el.style.display = 'none'; });
document.querySelectorAll('.tab').forEach(function(el) { el.classList.remove('active'); });
document.getElementById('tab-' + name).style.display = 'block';
this.classList.add('active');
});
});
</script>
</body>
</html>"""

EDIT_HTML = """<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Edit Entry - QUANTAM-JWIZ-INVENTORY</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Inter,-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;background:#0f172a;padding:30px 20px;color:#e2e8f0}
.container{max-width:700px;margin:0 auto}
h1{font-size:22px;font-weight:700;color:#f1f5f9;margin-bottom:4px}
.sub{color:#94a3b8;margin-bottom:24px;font-size:13px}
.card{background:#1e293b;border-radius:12px;padding:28px;border:1px solid #334155}
.form-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:14px}
.form-group{display:flex;flex-direction:column}
.form-group label{font-size:12px;font-weight:600;margin-bottom:6px;color:#94a3b8;text-transform:uppercase;letter-spacing:0.3px}
.form-group input{padding:10px 14px;border:1px solid #334155;border-radius:8px;font-size:14px;outline:0;transition:.2s;background:#0f172a;color:#e2e8f0}
.form-group input:focus{border-color:#60a5fa;box-shadow:0 0 0 3px rgba(96,165,250,.15)}
.btn{display:inline-flex;align-items:center;gap:6px;padding:10px 22px;border:none;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;transition:.2s;text-decoration:none}
.btn-primary{background:#3b82f6;color:#fff}
.btn-primary:hover{background:#2563eb}
.btn-secondary{background:#475569;color:#fff}
.btn-secondary:hover{background:#334155}
.actions{display:flex;gap:12px;margin-top:24px}
@media(max-width:600px){.form-grid{grid-template-columns:1fr}}
</style>
</head>
<body>
<div class="container">
<h1>Edit {{ typ|upper }} Entry</h1>
<p class="sub">Update the fields below</p>
<div class="card">
<form method="POST" action="/update/{{ typ }}/{{ idx }}">
<div class="form-grid">
{% for field in fields %}
{% if field != "DATE" %}
<div class="form-group">
<label for="f_{{ field }}">{{ field }}</label>
<input type="text" id="f_{{ field }}" name="{{ field }}" value="{{ row[field] }}" required>
</div>
{% else %}
<div class="form-group">
<label>{{ field }}</label>
<div style="padding:10px 14px;border:1px solid #334155;border-radius:8px;background:#0f172a;color:#94a3b8;font-size:14px">{{ row[field] }}</div>
</div>
{% endif %}
{% endfor %}
</div>
<div class="actions">
<button type="submit" class="btn btn-primary">Save Changes</button>
<a href="/" class="btn btn-secondary">Cancel</a>
</div>
</form>
</div>
</div>
</body>
</html>"""

@app.route("/")
def index():
    merged = request.args.get("merged")
    inbound = load_inbound()
    outbound = load_outbound()
    total_qty = sum(safe_int(r.get("NO OF UNIT", 0)) for r in inbound)
    unique_persons = len(set(r.get("EMPLOYEE NAME", "") for r in outbound if r.get("EMPLOYEE NAME", "").strip()))
    product_counts = {}
    for row in outbound:
        p = row.get("PRODUCT", "").strip().lower()
        q = safe_int(row.get("QUANTITY", 0))
        product_counts[p] = product_counts.get(p, 0) + q
    pc_count = product_counts.get("pc", 0)
    keyboard_count = product_counts.get("keyboard", 0)
    mouse_count = product_counts.get("mouse", 0)
    dept_qty = {}
    for row in outbound:
        d = row.get("DEPARTMENT", "").strip().upper()
        q = safe_int(row.get("QUANTITY", 0))
        if d and d != "IT":
            dept_qty[d] = dept_qty.get(d, 0) + q
    dept_items = sorted(dept_qty.items())
    return render_template_string(HTML,
        inbound=inbound,
        outbound=outbound,
        inbound_fields=INBOUND_HEADERS,
        outbound_fields=OUTBOUND_HEADERS,
        total_qty=total_qty,
        unique_persons=unique_persons,
        pc_count=pc_count,
        keyboard_count=keyboard_count,
        mouse_count=mouse_count,
        dept_items=dept_items,
        now=datetime.now().strftime("%Y-%m-%d %H:%M"),
        merged=merged)

@app.route("/add/inbound", methods=["POST"])
def add_inbound():
    today = datetime.now().strftime("%Y-%m-%d")
    row = {
        "PRODUCT": request.form.get("PRODUCT", ""),
        "SPECIFICATION": request.form.get("SPECIFICATION", ""),
        "NO OF UNIT": request.form.get("NO OF UNIT", "1"),
        "REMARKS": request.form.get("REMARKS", ""),
        "DATE": today
    }
    rows = load_inbound()
    merged = False
    for existing in rows:
        if (existing.get("PRODUCT", "") == row["PRODUCT"] and
            existing.get("SPECIFICATION", "") == row["SPECIFICATION"]):
            existing["NO OF UNIT"] = str(safe_int(existing.get("NO OF UNIT", 0)) + safe_int(row["NO OF UNIT"]))
            existing["DATE"] = row["DATE"]
            merged = True
            break
    if not merged:
        rows.append(row)
    save_inbound(rows)
    return redirect("/?merged=1" if merged else "/")

@app.route("/add/outbound", methods=["POST"])
def add_outbound():
    row = {
        "PRODUCT": request.form.get("PRODUCT", ""),
        "DEPARTMENT": request.form.get("DEPARTMENT", ""),
        "EMPLOYEE NAME": request.form.get("EMPLOYEE NAME", ""),
        "QUANTITY": request.form.get("QUANTITY", "1"),
        "REMARKS": request.form.get("REMARKS", ""),
        "DATE": datetime.now().strftime("%Y-%m-%d")
    }
    rows = load_outbound()
    rows.append(row)
    save_outbound(rows)
    return redirect("/")

@app.route("/delete/inbound/<int:idx>")
def delete_inbound(idx):
    rows = load_inbound()
    if 0 <= idx < len(rows):
        rows.pop(idx)
        save_inbound(rows)
    return redirect("/")

@app.route("/delete/outbound/<int:idx>")
def delete_outbound(idx):
    rows = load_outbound()
    if 0 <= idx < len(rows):
        rows.pop(idx)
        save_outbound(rows)
    return redirect("/")

@app.route("/edit/<typ>/<int:idx>")
def edit_entry(typ, idx):
    if typ == "inbound":
        rows = load_inbound()
        fields = INBOUND_HEADERS
    elif typ == "outbound":
        rows = load_outbound()
        fields = OUTBOUND_HEADERS
    else:
        return redirect("/")
    if idx < 0 or idx >= len(rows):
        return redirect("/")
    return render_template_string(EDIT_HTML, typ=typ, idx=idx, row=rows[idx], fields=fields)

@app.route("/update/<typ>/<int:idx>", methods=["POST"])
def update_entry(typ, idx):
    if typ == "inbound":
        rows = load_inbound()
        headers = INBOUND_HEADERS
    elif typ == "outbound":
        rows = load_outbound()
        headers = OUTBOUND_HEADERS
    else:
        return redirect("/")
    if idx < 0 or idx >= len(rows):
        return redirect("/")
    today = datetime.now().strftime("%Y-%m-%d")
    row = {h: request.form.get(h, rows[idx].get(h, "")) for h in headers}
    row["DATE"] = rows[idx].get("DATE", today)
    rows[idx] = row
    if typ == "inbound":
        save_inbound(rows)
    else:
        save_outbound(rows)
    return redirect("/")

def normalize_col(name):
    return " ".join(name.strip().upper().split())

def safe_int(val, default=0):
    try:
        return int(val)
    except (ValueError, TypeError):
        return default

def col_match(name, target):
    return normalize_col(name) == normalize_col(target)

@app.route("/import/<typ>", methods=["POST"])
def import_excel(typ):
    if typ not in ("inbound", "outbound"):
        return redirect("/")
    if openpyxl is None:
        return "openpyxl not installed", 500
    file = request.files.get("file")
    if not file:
        return redirect("/")
    wb = openpyxl.load_workbook(file)
    sheet_name = request.form.get("sheet", "").strip()
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            return f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}", 400
        ws = wb[sheet_name]
    else:
        ws = wb.active
    headers = []
    for cell in ws[1]:
        h = cell.value.strip() if cell.value else ""
        headers.append(h)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                d[headers[i]] = str(val) if val is not None else ""
        if any(v.strip() for v in d.values()):
            rows.append(d)
    if not rows:
        return redirect("/")

    today = datetime.now().strftime("%Y-%m-%d")
    target_headers = INBOUND_HEADERS if typ == "inbound" else OUTBOUND_HEADERS
    col_map = {}
    for h in headers:
        for th in target_headers:
            if col_match(h, th):
                col_map[h] = th
                break

    if typ == "inbound":
        existing = load_inbound()
        for row in rows:
            mapped = {}
            for src, tgt in col_map.items():
                mapped[tgt] = row.get(src, "")
            mapped.setdefault("NO OF UNIT", "1")
            mapped["DATE"] = today
            merged = False
            for er in existing:
                if (er.get("PRODUCT", "") == mapped.get("PRODUCT", "") and
                    er.get("SPECIFICATION", "") == mapped.get("SPECIFICATION", "")):
                    er["NO OF UNIT"] = str(safe_int(er.get("NO OF UNIT", 0)) + safe_int(mapped.get("NO OF UNIT", 1)))
                    er["DATE"] = today
                    merged = True
                    break
            if not merged:
                existing.append(mapped)
        save_inbound(existing)
    else:
        existing = load_outbound()
        for row in rows:
            mapped = {}
            for src, tgt in col_map.items():
                mapped[tgt] = row.get(src, "")
            mapped["DATE"] = today
            existing.append(mapped)
        save_outbound(existing)

    return redirect("/")

@app.route("/export/csv")
def export_csv():
    ib = load_inbound()
    ob = load_outbound()
    if not ib and not ob:
        return "No data to export", 400
    combined_headers = ["TYPE"] + OUTBOUND_HEADERS
    si = io.StringIO()
    w = csv.DictWriter(si, fieldnames=combined_headers)
    w.writeheader()
    for r in ib:
        row = {"TYPE": "INBOUND"}
        row.update({k: r.get(k, "") for k in OUTBOUND_HEADERS})
        w.writerow(row)
    for r in ob:
        row = {"TYPE": "OUTBOUND"}
        row.update({k: r.get(k, "") for k in OUTBOUND_HEADERS})
        w.writerow(row)
    mem = io.BytesIO(si.getvalue().encode("utf-8-sig"))
    return send_file(mem, mimetype="text/csv", as_attachment=True, download_name=f"inventory_{datetime.now().strftime('%Y%m%d')}.csv")

@app.route("/export/excel")
def export_excel():
    if openpyxl is None:
        return "openpyxl not installed", 500
    inbound = load_inbound()
    outbound = load_outbound()
    if not outbound and not inbound:
        return "No data to export", 400
    wb = openpyxl.Workbook()
    bold_font = Font(bold=True, color="FFFFFF")
    fill = openpyxl.styles.PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
    align = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    def write_sheet(ws, headers, rows, title):
        ws.title = title
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = bold_font
            c.fill = fill
            c.alignment = align
            c.border = thin
        for ri, row in enumerate(rows, 2):
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=ri, column=ci, value=row.get(h, ""))
                c.border = thin
                c.alignment = Alignment(horizontal="center", vertical="center")
        for ci in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 22

    ws1 = wb.active
    write_sheet(ws1, ["S.NO"] + INBOUND_HEADERS, inbound, "Inbound")
    if outbound:
        ws2 = wb.create_sheet()
        write_sheet(ws2, ["S.NO"] + OUTBOUND_HEADERS, outbound, "Outbound")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"inventory_{datetime.now().strftime('%Y%m%d')}.xlsx")

@app.route("/export/pdf")
def export_pdf():
    if FPDF is None:
        return "fpdf2 not installed", 500
    inbound = load_inbound()
    outbound = load_outbound()
    if not outbound and not inbound:
        return "No data to export", 400
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 12, "Inventory Report", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 8, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    def draw_table(headers, rows):
        w_avail = 277
        col_w = w_avail / len(headers)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(67, 97, 238)
        pdf.set_text_color(255, 255, 255)
        for h in headers:
            pdf.cell(col_w, 8, h, border=1, align="C", fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 8)
        for row in rows:
            for h in headers:
                pdf.cell(col_w, 7, row.get(h, "")[:30], border=1, align="C")
            pdf.ln()

    if inbound:
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 10, "Inbound - Stock In", new_x="LMARGIN", new_y="NEXT")
        draw_table(["S.NO"] + INBOUND_HEADERS, [{**{"S.NO": str(i+1)}, **r} for i, r in enumerate(inbound)])
        pdf.ln(5)
    if outbound:
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 10, "Outbound - Issued to Person", new_x="LMARGIN", new_y="NEXT")
        draw_table(["S.NO"] + OUTBOUND_HEADERS, [{**{"S.NO": str(i+1)}, **r} for i, r in enumerate(outbound)])
    buf = io.BytesIO(pdf.output())
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=f"inventory_{datetime.now().strftime('%Y%m%d')}.pdf")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
